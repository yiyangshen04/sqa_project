from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from uat_data import TEST_CASES, USER_STORIES

COLUMNS = [
    ("id", "Test Case ID", 12),
    ("name", "Name", 45),
    ("story", "User Story", 14),
    ("description", "Description", 50),
    ("preconditions", "Preconditions", 35),
    ("steps", "Steps", 55),
    ("expected", "Expected Result", 50),
    ("actual", "Actual Result", 30),
    ("result", "Pass/Fail", 12),
    ("technique", "Black-box Technique", 25),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Test Cases"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, case in enumerate(TEST_CASES, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case[key])
            cell.alignment = wrap

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("User Stories")
    for idx, header in enumerate(["Story ID", "User Story"], start=1):
        cell = ws2.cell(row=1, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 90
    for row_idx, (story_id, text) in enumerate(USER_STORIES, start=2):
        ws2.cell(row=row_idx, column=1, value=story_id)
        ws2.cell(row=row_idx, column=2, value=text).alignment = wrap

    return wb


def main() -> None:
    out = Path(__file__).parent / "uat_test_cases.xlsx"
    wb = build_workbook()
    wb.save(out)
    print(f"wrote {out} with {len(TEST_CASES)} test cases")


if __name__ == "__main__":
    main()
